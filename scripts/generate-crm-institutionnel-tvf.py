from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

ROOT = Path(r'C:\Users\jowst\Documents\TERRITOIRES VIVANTS FRANCE')
OUT = ROOT / 'documents' / 'crm-institutionnel-tvf'
OUT.mkdir(parents=True, exist_ok=True)
path = OUT / 'registre-crm-institutionnel-tvf-papier.xlsx'

GREEN = '18392F'
SAGE = 'EAF2ED'
GOLD = 'B4945A'
PALE = 'F6F8F6'
WHITE = 'FFFFFF'
TEXT = '1F2937'
GRAY = '667085'
BORDER = 'D9E0DC'
BLUE = 'E8F1F5'
BEIGE = 'F4EFE6'

headers = [
    'Priorité', 'Famille', 'Organisme / ministère / programme', 'Responsable institutionnel à viser',
    'Nom connu / à compléter', 'Lien avec l’objet de TVF', 'Angle de prise de contact', 'Document TVF à envoyer',
    'Site / source officielle', 'E-mail / formulaire', 'Téléphone', 'Adresse courrier', 'Date 1er contact',
    'Canal utilisé', 'Personne contactée', 'Réponse reçue', 'Date relance', 'Statut', 'Prochaine action', 'Notes'
]

status_options = ['À contacter', 'Contacté', 'Relance à faire', 'Réponse reçue', 'RDV demandé', 'RDV obtenu', 'Dossier envoyé', 'À classer', 'Sans suite']
priority_options = ['Haute', 'Moyenne', 'À suivre']
channel_options = ['Courrier', 'E-mail', 'Formulaire', 'Téléphone', 'Rendez-vous', 'LinkedIn', 'Autre']

def row(prio, fam, org, resp, name, link, angle, doc, site):
    return [prio, fam, org, resp, name, link, angle, doc, site, '', '', '', '', '', '', '', '', 'À contacter', '', '']

ministeres = [
    row('Haute','Ministère','Ministère de la Ville et du Logement','Ministre / cabinet / direction logement','Vincent Jeanbrun','Logements vacants, habitat privé, rénovation, propriétaires','Présenter TVF comme outil de repérage, qualification et remise en usage des logements vacants','Courrier ministère Ville et Logement + dossier TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('Haute','Ministère','Ministère de l’Aménagement du territoire et de la Décentralisation','Ministre / cabinet / direction territoires','Françoise Gatel','Collectivités, ingénierie locale, territoires pilotes, revitalisation','Demander une mise en relation avec services compétents pour expérimentation territoriale','Courrier Aménagement du territoire + dossier TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('Haute','Ministère','Ministère de la Transition écologique, Biodiversité et Climat','Ministre / cabinet / directions transition écologique','Monique Barbut','Réemploi, friches, sobriété foncière, matériaux, bâtiment durable','Présenter Matériauthèque, réemploi et remise en usage du patrimoine existant','Courrier Transition écologique + brochure matériaux','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('Haute','Ministère','Ministère des PME, du Commerce, de l’Artisanat, du Tourisme et du Pouvoir d’achat','Ministre / cabinet / direction commerce artisanat','Serge Papin','Commerces vacants, artisans, centralités, entreprises partenaires','Présenter Commerce Vivant et démarche locaux commerciaux vacants','Courrier PME Commerce + brochure entreprises','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('Moyenne','Ministère','Ministère de l’Économie, des Finances, Souveraineté industrielle, énergétique et numérique','Ministre / cabinet / directions économie et numérique','Roland Lescure','Financements, entreprises, innovation, outils numériques TVF OS / TVF Mobile','Présenter TVF comme outil d’organisation économique des ressources dormantes','Courrier Économie Finances + dossier TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('Moyenne','Ministère','Ministère du Travail et des Solidarités','Ministre / cabinet / directions insertion et solidarités','Jean-Pierre Farandou','Insertion, chantiers encadrés, utilité sociale, accompagnement des publics','Présenter pôle Solidarité et Insertion lié aux projets immobiliers utiles','Courrier Travail Solidarités + brochure pôle Solidarité','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('Moyenne','Ministère','Ministère des Sports, de la Jeunesse et de la Vie associative','Ministre / cabinet / direction vie associative','Marina Ferrari','Structuration associative, bénévolat, mobilisation citoyenne','Présenter TVF comme association structurée au service des territoires','Courrier Vie associative + dossier TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('À suivre','Ministère','Ministère de la Culture','Ministre / cabinet / direction patrimoine','Catherine Pégard','Patrimoine bâti, centres anciens, bâtiments délaissés','Présenter TVF sur patrimoine vacant et usages culturels/associatifs possibles','Courrier Culture + dossier TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('À suivre','Ministère','Ministère de la Justice','Ministre / cabinet / direction affaires civiles','Gérald Darmanin','Successions, indivisions, biens sans maître, cadre juridique','Demander bonnes pratiques et orientation sur situations juridiques sensibles','Courrier Justice + note juridique TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
    row('À suivre','Ministère','Ministère de l’Agriculture et de la Souveraineté alimentaire','Ministre / cabinet / services ruralité/foncier','Annie Genevard','Bâtiments ruraux vacants, foncier rural, villages, revitalisation rurale','Présenter adaptation future de la méthode TVF aux territoires ruraux','Courrier Agriculture Ruralité + dossier TVF','https://www.info.gouv.fr/composition-du-gouvernement'),
]

agences = [
    row('Haute','Agence nationale','ANCT - Agence nationale de la cohésion des territoires','Direction générale / programmes territoires','À compléter','Revitalisation, Action Cœur de Ville, Petites Villes de Demain, Villages d’avenir','Demander orientation vers programmes de revitalisation territoriale','Dossier TVF + courrier Aménagement du territoire','https://agence-cohesion-territoires.gouv.fr/'),
    row('Haute','Agence nationale','Anah - Agence nationale de l’habitat','Direction générale / délégation locale Anah','À compléter','Habitat privé, rénovation, propriétaires, logements vacants','Présenter parcours propriétaire et qualification des biens vacants','Brochure propriétaires + dossier TVF','https://www.anah.gouv.fr/'),
    row('Haute','Agence nationale','ADEME','Direction générale / directions régionales','À compléter','Réemploi, économie circulaire, bâtiment durable, déchets de chantier','Présenter Matériauthèque Solidaire et logique ressources dormantes','Brochure matériaux + courrier Transition écologique','https://www.ademe.fr/'),
    row('Haute','Agence nationale','Banque des Territoires / Caisse des Dépôts','Direction régionale / responsable développement territorial','À compléter','Financement territorial, revitalisation, foncier, numérique d’intérêt général','Demander échange sur financement expérimentation et TVF OS','Dossier TVF + programme territoire pilote','https://www.banquedesterritoires.fr/'),
    row('Moyenne','Agence nationale','Cerema','Direction territoriale / expertise bâtiment et territoire','À compléter','Expertise technique, territoires, foncier, observation, mobilité, bâtiment','Demander appui méthodologique ou orientation sur observation territoriale','Dossier TVF + note Observatoire','https://www.cerema.fr/'),
    row('Moyenne','Agence nationale','ANRU','Direction générale / programmes quartiers','À compléter','Renouvellement urbain, quartiers, habitat, commerces, équipements','Identifier passerelles avec quartiers en renouvellement urbain','Dossier TVF + courrier institutionnel','https://www.anru.fr/'),
    row('Moyenne','Agence nationale','Action Logement','Direction régionale / responsable partenariats','À compléter','Logement, emploi, salariés, rénovation, mobilisation logement','Présenter remise en usage logements pour besoins territoriaux','Brochure propriétaires + dossier TVF','https://www.actionlogement.fr/'),
    row('Moyenne','Agence nationale','Bpifrance','Direction régionale / innovation et entrepreneuriat','À compléter','Entreprises, innovation, outils numériques, partenariats économiques','Présenter TVF OS / TVF Mobile comme innovation territoriale','Dossier TVF + note outils numériques','https://www.bpifrance.fr/'),
    row('À suivre','Agence nationale','France Active','Direction régionale / financement ESS','À compléter','Financement associatif, économie sociale et solidaire','Étudier financement structuration TVF et projets territoriaux','Dossier TVF + budget prévisionnel','https://www.franceactive.org/'),
    row('À suivre','Réseau national','Union sociale pour l’habitat','Direction régionale / bailleurs sociaux','À compléter','Logement social, patrimoine, remise en usage, partenariats','Identifier passerelles avec bailleurs et logements accompagnés','Dossier TVF + note logement','https://www.union-habitat.org/'),
    row('À suivre','Réseau national','SOLIHA','Fédération / association locale','À compléter','Habitat privé, accompagnement propriétaires, rénovation sociale','Demander complémentarité et orientation locale','Brochure propriétaires + courrier logement','https://soliha.fr/'),
    row('À suivre','Réseau national','CCI France / CCI territoriale','Présidence / direction développement économique','À compléter','Commerces, entreprises, locaux vacants, artisans','Repérer partenaires entreprises et locaux commerciaux','Brochure entreprises + Commerce Vivant','https://www.cci.fr/'),
    row('À suivre','Réseau national','CMA France / Chambre des métiers locale','Direction / développement artisanal','À compléter','Artisans, rénovation, compétences, chantiers, commerces','Mobiliser artisans et entreprises de rénovation','Brochure entreprises + matériaux','https://www.artisanat.fr/'),
]

programmes = [
    row('Haute','Programme national','Action Cœur de Ville','ANCT / Banque des Territoires / collectivités concernées','Responsable programme local à identifier','Centres-villes, commerces, habitat, revitalisation','Positionner TVF comme outil de repérage et suivi des biens vacants','Dossier TVF + note Commerce/Habitat','https://agence-cohesion-territoires.gouv.fr/action-coeur-de-ville'),
    row('Haute','Programme national','Petites Villes de Demain','ANCT / chef de projet local','Responsable programme local à identifier','Bourgs, centralités, ingénierie, habitat et commerce','Proposer méthode TVF comme appui opérationnel local','Dossier TVF + note méthode','https://agence-cohesion-territoires.gouv.fr/petites-villes-de-demain'),
    row('Haute','Programme national','Villages d’avenir','ANCT / préfecture / chef de projet','Responsable programme local à identifier','Ruralité, ingénierie, projets communaux, bâtiments délaissés','Étudier adaptation TVF aux communes rurales','Dossier TVF + courrier ruralité','https://agence-cohesion-territoires.gouv.fr/'),
    row('Haute','Programme national','France Ruralités','ANCT / préfectures / collectivités','Responsable local à identifier','Ruralité, revitalisation, services, bâti vacant','Identifier cadre de déploiement hors métropole','Dossier TVF + note territoire pilote','https://agence-cohesion-territoires.gouv.fr/france-ruralites'),
    row('Haute','Programme national','Opération de Revitalisation de Territoire - ORT','Collectivité / État / ANCT','Responsable ORT local à identifier','Habitat, commerces, centralités, foncier, revitalisation','Faire entrer TVF comme outil de suivi et repérage si conventionné','Dossier TVF + courrier collectivité','https://www.ecologie.gouv.fr/'),
    row('Haute','Programme national','OPAH / OPAH-RU','Collectivité / Anah / opérateur habitat','Chef de projet OPAH à identifier','Habitat privé ancien, propriétaires, rénovation','Complémentarité avec repérage propriétaires et biens vacants','Brochure propriétaires + note habitat','https://www.anah.gouv.fr/'),
    row('Moyenne','Programme national','PIG - Programme d’intérêt général habitat','Collectivité / Anah','Responsable PIG à identifier','Habitat privé, rénovation, accompagnement propriétaires','Orienter propriétaires vers dispositifs existants','Brochure propriétaires','https://www.anah.gouv.fr/'),
    row('Moyenne','Programme national','MaPrimeRénov’','Anah / France Rénov’','Guichet France Rénov local','Rénovation énergétique des logements','Orienter propriétaires vers guichets habilités, sans promettre d’aide','Brochure propriétaires','https://www.maprimerenov.gouv.fr/'),
    row('Moyenne','Programme national','France Rénov’','Anah / ADEME / collectivités','Espace conseil local','Conseil rénovation, accompagnement propriétaires','Trouver guichet local pour propriétaires accompagnés','Brochure propriétaires','https://france-renov.gouv.fr/'),
    row('Moyenne','Programme national','CEE - Certificats d’économies d’énergie','Ministère énergie / obligés / délégataires','Interlocuteur CEE à identifier','Financement rénovation énergétique','Identifier possibilités selon projets, sans engagement automatique','Note financement','https://www.ecologie.gouv.fr/'),
    row('Moyenne','Programme national','Fonds vert','Préfecture / DREAL / collectivités','Référent Fonds vert local','Transition écologique, friches, rénovation, adaptation','Identifier appels à projets mobilisables par collectivités','Dossier TVF + note transition','https://www.ecologie.gouv.fr/'),
    row('Moyenne','Programme national','Fonds friches / recyclage foncier','Préfecture / DREAL / collectivités','Référent friches local','Friches, foncier, recyclage urbain','Positionner TVF sur repérage/qualification avant opérateur technique','Note friches','https://www.ecologie.gouv.fr/'),
    row('Moyenne','Programme national','Zéro Artificialisation Nette - ZAN','Ministère transition / collectivités','Service urbanisme / planification','Sobriété foncière, réutilisation patrimoine existant','Démontrer lien entre remise en usage et limitation extension urbaine','Note Observatoire','https://www.ecologie.gouv.fr/'),
    row('À suivre','Programme national','Territoires d’industrie','ANCT / collectivités / Banque des Territoires','Chef de projet local à identifier','Friches économiques, bâtiments d’activité, entreprises','Identifier passerelles pour locaux et sites d’activité vacants','Dossier TVF + note entreprises','https://www.territoires-industrie.gouv.fr/'),
    row('À suivre','Programme national','Quartiers 2030 / Politique de la ville','ANCT / préfecture / collectivités','Chef de projet politique de la ville','Quartiers, cadre de vie, locaux, insertion','Lier remise en usage, solidarité, commerce et projets locaux','Dossier TVF + pôle Solidarité','https://agence-cohesion-territoires.gouv.fr/'),
]

epf = [
    row('Haute','Établissement foncier','EPORA - Établissement public foncier de l’Ouest Rhône-Alpes','Direction / service territoire Loire','À compléter','Foncier, friches, bâtiments délaissés, Saint-Étienne / Loire','Contact prioritaire pour foncier et friches du territoire pilote','Courrier EPORA + dossier territoire pilote','https://www.epora.fr/'),
    row('Haute','Établissement foncier','EPF Auvergne-Rhône-Alpes / réseau foncier régional','Direction / antenne régionale','À vérifier','Foncier régional, recyclage, portage foncier','Identifier compétence territoriale exacte autour de Saint-Étienne','Dossier TVF + note foncier','https://www.epf-ara.fr/'),
    row('Moyenne','Établissement foncier','EPF d’Île-de-France','Direction partenariats','À compléter','Méthode foncière, friches, portage, recyclage urbain','Contact benchmark et éventuel futur déploiement','Dossier TVF','https://www.epfif.fr/'),
    row('Moyenne','Établissement foncier','EPF Provence-Alpes-Côte d’Azur','Direction territoriale','À compléter','Foncier, habitat, renouvellement urbain','Contact benchmark / futur partenariat territorial','Dossier TVF','https://www.epfprovencealpescotedazur.fr/'),
    row('Moyenne','Établissement foncier','EPF Hauts-de-France','Direction territoriale','À compléter','Recyclage foncier, friches, habitat, centralités','Contact benchmark et méthode friches','Dossier TVF','https://www.epf-hdf.fr/'),
    row('À suivre','Établissement foncier','EPF Normandie','Direction territoriale','À compléter','Foncier, friches, revitalisation','Contact futur selon expansion TVF','Dossier TVF','https://www.epf-normandie.fr/'),
    row('À suivre','Établissement foncier','EPF Nouvelle-Aquitaine','Direction territoriale','À compléter','Foncier, collectivités, sobriété foncière','Contact futur selon expansion TVF','Dossier TVF','https://www.epfna.fr/'),
    row('À suivre','Établissement foncier','EPF Grand Est','Direction territoriale','À compléter','Foncier, friches, renouvellement urbain','Contact futur selon expansion TVF','Dossier TVF','https://www.epfge.fr/'),
    row('À suivre','Établissement foncier','EPF Bretagne','Direction territoriale','À compléter','Foncier, sobriété, collectivités','Contact futur selon expansion TVF','Dossier TVF','https://www.epfbretagne.fr/'),
    row('À suivre','Établissement foncier','EPF Occitanie','Direction territoriale','À compléter','Foncier, habitat, friches','Contact futur selon expansion TVF','Dossier TVF','https://www.epf-occitanie.fr/'),
]

acteurs = [
    row('Haute','Collectivité / État local','Préfecture de la Loire','Préfet / Secrétaire général / DDT','À compléter','Coordination État local, sécurité, logement, friches, dispositifs','Présenter territoire pilote et demander orientation services','Courrier Préfecture + dossier TVF','https://www.loire.gouv.fr/'),
    row('Haute','Collectivité','Ville de Saint-Étienne','Maire / adjoints habitat, commerce, urbanisme','À compléter','Territoire pilote, biens vacants, commerces, friches','Construire cadre local d’expérimentation','Courrier Ville + programme pilote','https://www.saint-etienne.fr/'),
    row('Haute','Collectivité','Saint-Étienne Métropole','Président / services habitat, économie, foncier','À compléter','Intercommunalité, habitat, développement économique, foncier','Proposer démarche métropolitaine autour de la vacance','Courrier Métropole + dossier TVF','https://www.saint-etienne-metropole.fr/'),
    row('Moyenne','Collectivité','Département de la Loire','Président / services insertion, habitat, solidarité','À compléter','Insertion, solidarités, logement, territoires','Présenter lien entre revitalisation et utilité sociale','Courrier Département + pôle Solidarité','https://www.loire.fr/'),
    row('Moyenne','Collectivité','Région Auvergne-Rhône-Alpes','Président / services aménagement, économie, formation','À compléter','Formation, économie, aménagement, entreprises','Identifier aides régionales et réseaux entreprises','Dossier TVF','https://www.auvergnerhonealpes.fr/'),
    row('Moyenne','Organisme local','CCI Lyon Métropole Saint-Étienne Roanne','Direction territoriale Saint-Étienne','À compléter','Commerces, entreprises, locaux vacants','Mobiliser entreprises, artisans, commerces','Brochure entreprises','https://www.lyon-metropole.cci.fr/'),
    row('Moyenne','Organisme local','CAPEB Loire','Présidence / direction','À compléter','Artisans du bâtiment, rénovation, chantiers','Partenariat artisans et travaux encadrés','Courrier CAPEB + brochure entreprises','https://www.capeb.fr/'),
    row('À suivre','Organisme local','Maison de l’Emploi / PLIE Saint-Étienne','Direction / référent insertion','À compléter','Insertion, emploi, publics accompagnés','Créer passerelles chantiers encadrés et formation','Courrier insertion + pôle Solidarité','À compléter'),
    row('À suivre','Organisme local','Mission Locale de Saint-Étienne','Direction / référent entreprises','À compléter','Jeunes, insertion, découverte métiers','Proposer actions découverte métiers et chantiers encadrés','Courrier insertion','https://www.missionlocale-st-etienne.org/'),
    row('À suivre','Réseau ESS','GESAT','Direction partenariats','À compléter','EA/ESAT, achats responsables, insertion','Identifier acteurs adaptés pour services et chantiers','Dossier TVF + pôle Solidarité','https://www.reseau-gesat.com/'),
    row('À suivre','Financeur / fondation','Fondation de France','Programme habitat / territoires / solidarités','À compléter','Financement associatif et projets territoriaux','Demander orientation vers appel à projets adapté','Courrier fondation + dossier TVF','https://www.fondationdefrance.org/'),
    row('À suivre','Fondation entreprise','Fondation VINCI pour la Cité','Délégué régional / correspondant local','À compléter','Insertion, chantiers, mécénat de compétences','Présenter projets utiles et besoin mécénat','Courrier fondation + dossier TVF','https://www.fondation-vinci.com/'),
    row('À suivre','Fondation entreprise','Fondation Bouygues / Bouygues Construction','Fondation / RSE territoire','À compléter','Bâtiment, matériaux, solidarité, territoires','Identifier mécénat, matériaux ou expertise','Courrier fondation + brochure entreprises','https://www.bouygues-construction.com/'),
    row('À suivre','Fondation entreprise','Fondation EDF','Fondation / solidarité / environnement','À compléter','Solidarité, transition, projets locaux','Demander orientation financement/mécénat','Courrier fondation','https://fondation.edf.com/'),
    row('À suivre','Fondation banque','Fondation Crédit Agricole / Fondation CA Solidarité Développement','Correspondant régional','À compléter','Territoires, inclusion, patrimoine, logement','Présenter projet territorial pilote','Courrier fondation','https://www.credit-agricole.com/'),
]

blank_rows = [[ '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'À contacter', '', '' ] for _ in range(30)]

wb = Workbook()
ws = wb.active
ws.title = 'Mode emploi'

thin = Side(style='thin', color=BORDER)
medium = Side(style='medium', color=GREEN)

for sheet in wb.worksheets:
    pass

def setup_sheet(ws, title, subtitle, rows):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15)
    ws.freeze_panes = 'A5'
    ws.print_title_rows = '1:4'
    ws['A1'] = 'TERRITOIRES VIVANTS FRANCE'
    ws['A1'].font = Font(name='Manrope', size=16, bold=True, color=GREEN)
    ws['A2'] = title
    ws['A2'].font = Font(name='Manrope', size=13, bold=True, color=TEXT)
    ws['A3'] = subtitle
    ws['A3'].font = Font(name='Inter', size=10, color=GRAY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    for col, header in enumerate(headers, 1):
        c = ws.cell(4, col, header)
        c.font = Font(name='Inter', size=9, bold=True, color=WHITE)
        c.fill = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(top=medium, bottom=medium, left=thin, right=thin)
    for r_idx, data in enumerate(rows, 5):
        for c_idx, val in enumerate(data, 1):
            c = ws.cell(r_idx, c_idx, val)
            c.font = Font(name='Inter', size=8.2, color=TEXT)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if r_idx % 2 == 0:
                c.fill = PatternFill('solid', fgColor='FBFCFA')
        ws.row_dimensions[r_idx].height = 48
    widths = [9,15,34,28,22,40,42,30,34,24,17,28,15,14,22,22,15,16,30,36]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.auto_filter.ref = f'A4:T{max(5, 4+len(rows))}'
    ws.print_area = f'A1:T{max(5, 4+len(rows))}'
    # validations
    dv_status = DataValidation(type='list', formula1='"' + ','.join(status_options) + '"', allow_blank=True)
    dv_prio = DataValidation(type='list', formula1='"' + ','.join(priority_options) + '"', allow_blank=True)
    dv_channel = DataValidation(type='list', formula1='"' + ','.join(channel_options) + '"', allow_blank=True)
    ws.add_data_validation(dv_status); ws.add_data_validation(dv_prio); ws.add_data_validation(dv_channel)
    dv_prio.add(f'A5:A{4+len(rows)}')
    dv_channel.add(f'N5:N{4+len(rows)}')
    dv_status.add(f'R5:R{4+len(rows)}')

# Mode emploi
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws['A1'] = 'Registre CRM institutionnel TVF - version papier'
ws['A1'].font = Font(name='Manrope', size=18, bold=True, color=GREEN)
ws['A3'] = 'Objectif'
ws['A3'].font = Font(name='Manrope', size=13, bold=True, color=GREEN)
ws['A4'] = 'Ce fichier sert à imprimer un registre de prise de contact pour les ministères, agences nationales, programmes, établissements fonciers et acteurs utiles à Territoires Vivants France.'
ws['A6'] = 'Mode d’utilisation conseillé'
ws['A6'].font = Font(name='Manrope', size=13, bold=True, color=GREEN)
notes = [
    '1. Imprimer les onglets utiles en mode paysage, idéalement en A3 pour les tableaux complets.',
    '2. Commencer par les lignes en priorité Haute : logement, aménagement du territoire, transition écologique, commerce, ANCT, Anah, ADEME, Banque des Territoires, EPORA.',
    '3. Compléter à la main les colonnes Date 1er contact, Canal utilisé, Personne contactée, Réponse, Date relance et Prochaine action.',
    '4. Ne pas présenter un partenariat comme acquis tant qu’il n’est pas formalisé par écrit.',
    '5. Les noms de responsables non officiellement vérifiés restent volontairement “À compléter”.',
    '6. Utiliser les courriers ministériels TVF déjà générés pour les ministères correspondants.',
]
for i, note in enumerate(notes, 7):
    ws.cell(i,1,note).font = Font(name='Inter', size=11, color=TEXT)
    ws.cell(i,1).alignment = Alignment(wrap_text=True)
ws.column_dimensions['A'].width = 120
ws.print_area = 'A1:A14'

sheets = [
    ('Ministères', 'Ministères à contacter', 'Suivi papier des ministères directement liés à l’objet de TVF.', ministeres),
    ('Agences nationales', 'Agences nationales et réseaux à mobiliser', 'Acteurs nationaux utiles pour logement, rénovation, cohésion territoriale, réemploi, financement et expertise.', agences),
    ('Programmes nationaux', 'Programmes nationaux compatibles avec TVF', 'Programmes et dispositifs dans lesquels TVF peut s’inscrire sans promettre de financement automatique.', programmes),
    ('Etablissements fonciers', 'Établissements fonciers et acteurs fonciers', 'Contacts à suivre pour friches, portage foncier, recyclage urbain et sobriété foncière.', epf),
    ('Acteurs utiles', 'Acteurs locaux, financeurs et réseaux complémentaires', 'Contacts intéressants pour le territoire pilote, les entreprises, l’insertion, les fondations et les partenaires.', acteurs),
    ('Suivi vierge', 'Fiches vierges de suivi de contact', 'Pages vierges à imprimer pour ajouter de nouveaux contacts institutionnels.', blank_rows),
]
for name, title, subtitle, rows in sheets:
    ws = wb.create_sheet(name)
    setup_sheet(ws, title, subtitle, rows)

wb.save(path)
print(path)

